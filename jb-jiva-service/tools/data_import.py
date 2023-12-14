from jugalbandi.legal_library import LegalKeys
from jugalbandi.library import (
  Library,
  LibraryFileType,
  DocumentMetaData,
  DocumentSupportingMetadata,
  DocumentFormat,
  Document
)
from jugalbandi.storage import GoogleStorage
from PIL import Image
from io import BytesIO
import httpx
import os
import fitz
import asyncio
import openpyxl
import json
import gspread
import re
import csv
from oauth2client.service_account import ServiceAccountCredentials
from datetime import datetime
from dotenv import load_dotenv
from jugalbandi.translator import GoogleTranslator
from jugalbandi.core.language import Language


# Function to get metadata from google sheets and convert it to csv file
def convert_google_sheets_meta_data_to_csv(required_sheet_name: str):
    credentials_path = os.environ["GOOGLE_APPLICATION_CREDENTIALS"]
    scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    creds = ServiceAccountCredentials.from_json_keyfile_name(credentials_path, scope)
    client = gspread.authorize(creds)
    sheet_url = os.environ["GOOGLE_SHEET_URL"]
    sheet = client.open_by_url(sheet_url)

    sheet_names = sheet.worksheets()
    for worksheet in sheet_names:
        print(worksheet.title)
        if worksheet.title == required_sheet_name:
            data = worksheet.get_all_values()
            keys = data[0]
            values = data[1:]
            meta_data_list = [dict(zip(keys, row)) for row in values]

            headers = meta_data_list[0].keys()
            csv_file = worksheet.title + ".csv"
            with open(csv_file, "w", newline="") as csv_output:
                writer = csv.DictWriter(csv_output, fieldnames=headers)
                writer.writeheader()
                for row in meta_data_list:
                    writer.writerow(row)


# Function to process raw metadata and set it in DocumentMetaData object for each document
async def set_meta_data(raw_meta_data: dict) -> DocumentMetaData:
    def capitalize_words(match):
        return match.group(0).capitalize()

    raw_doc_title = raw_meta_data["Doc title"]
    doc_title = re.sub(r'\b\w+\b', capitalize_words, raw_doc_title)

    if raw_meta_data["Publish Date"] != "":
        publish_date = datetime.strptime(raw_meta_data["Publish Date"], "%d/%m/%Y")
        formatted_publish_date = publish_date.strftime("%Y-%m-%d")
    else:
        formatted_publish_date = None

    if raw_meta_data["Pass Date"] != "":
        pass_date = datetime.strptime(raw_meta_data["Pass Date"], "%d/%m/%Y")
        formatted_pass_date = pass_date.strftime("%Y-%m-%d")
    else:
        formatted_pass_date = ""

    if raw_meta_data["Effective From"] != "":
        effective_from_date = datetime.strptime(raw_meta_data["Effective From"], "%d/%m/%Y")
        formatted_effective_from_date = effective_from_date.strftime("%Y-%m-%d")
    else:
        formatted_effective_from_date = ""

    doc_type = raw_meta_data["Type"].lower().strip()
    if (doc_type != "act" and raw_meta_data["Related Act title"] != "NA" and
       raw_meta_data["Related Act no"] != "NA" and raw_meta_data["Related Act Year"] != "NA"):
        act_title = raw_meta_data["Related Act title"]
        act_no = raw_meta_data["Related Act no"]
        act_year = raw_meta_data["Related Act Year"]
    else:
        act_no = raw_meta_data["Act no"]
        act_year = raw_meta_data["Act Year"]
        act_title = doc_title

    document_meta_data = DocumentMetaData(title=doc_title,
                                          original_file_name=raw_meta_data["File Name"],
                                          original_format=DocumentFormat.PDF,
                                          publish_date=formatted_publish_date,
                                          extra_data={
                                            LegalKeys.LEGAL_ACT_JURISDICTION.value: raw_meta_data["Jurisdiction"].lower().strip(),
                                            LegalKeys.LEGAL_ACT_NO.value: act_no,
                                            LegalKeys.LEGAL_ACT_YEAR.value: act_year,
                                            LegalKeys.LEGAL_ACT_TITLE.value: act_title,
                                            LegalKeys.LEGAL_DOC_TYPE.value: doc_type,
                                            LegalKeys.LEGAL_MINISTRY.value: raw_meta_data["Ministry"].strip(),
                                            LegalKeys.LEGAL_PASS_DATE.value: formatted_pass_date,
                                            LegalKeys.LEGAL_EFFECTIVE_DATE.value: formatted_effective_from_date
                                            })
    return document_meta_data


# Function to upload bare pdf act document to cloud storage and return Document object
async def upload_file(jiva_library: Library, document_file_path: str, document_meta_data: DocumentMetaData):
    file = open(document_file_path, "rb")
    file_bytes_content = file.read()
    file.close()
    document: Document = await jiva_library.add_document(document_meta_data, file_bytes_content)
    await document.make_public()
    return document


# Function to upload thumbnail image (1st page of act) to cloud storage
async def upload_thumbnail(document: Document):
    document_meta_data = await document.read_metadata()
    response = httpx.get(document_meta_data.public_url)
    buffer = BytesIO(response.content)
    pdf_document = fitz.open(stream=buffer.read(), filetype="pdf")
    page = pdf_document.load_page(0)
    pix = page.get_pixmap()
    print("AFTER pixmap")
    image = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    image_buffer = BytesIO()
    image.save(image_buffer, format="PNG")
    image_buffer.seek(0)
    document_supporting_metadata = DocumentSupportingMetadata(doc_id=document.id,
                                                              name="thumbnail.png",
                                                              original_file_name=document_meta_data.original_file_name,
                                                              extra_data={})
    document_meta_data = await document.write_supporting_document(document_supporting_metadata,
                                                                  "thumbnail.png",
                                                                  image_buffer.read())
    await document.make_public(file_path="thumbnail.png",
                               file_type=LibraryFileType.SUPPORTING)


# Function to upload sections json file to cloud storage for each document
async def upload_section(document: Document):
    document_meta_data = await document.read_metadata()
    original_file_name = document_meta_data.original_file_name
    print("Original File Name:", original_file_name)
    excel_file_path = os.environ["SECTIONS_EXCEL_PATH"]
    workbook = openpyxl.load_workbook(excel_file_path)

    section_list = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        first_cell_value = sheet.cell(row=1, column=1).value
        if first_cell_value == original_file_name:
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if row_index == 2:
                    keys = row
                elif row_index > 2:
                    section_data = {}
                    for col_index, cell_value in enumerate(row):
                        section_data[keys[col_index]] = str(cell_value).strip()
                    section_list.append(section_data)

    json_data = json.dumps(section_list)
    section_bytes_data = json_data.encode('utf-8')
    await document.write_sections(section_bytes_data)


# Function to upload all documents along with their sections & metadata in given csv file to cloud storage
async def act_uploading_process(jiva_library: Library, csv_file_name: str):
    raw_meta_data = []
    with open(csv_file_name, "r") as csv_input:
        reader = csv.DictReader(csv_input)
        for row in reader:
            raw_meta_data.append(dict(row))

    counter = 1
    base_act_path = os.environ["ACTS_PATH"]
    for meta_data in raw_meta_data:
        file_path = os.path.join(base_act_path, meta_data["File Name"])
        document_meta_data = await set_meta_data(meta_data)
        document = await upload_file(jiva_library, file_path, document_meta_data)
        print("\nFile Count:", counter)
        print("Document ID:", document.id)
        print("Document Title:", document_meta_data.title)
        await upload_thumbnail(document)
        await upload_section(document)
        with open("tools/docs_meta_data.csv", "a", newline="") as csv_output:
            writer = csv.DictWriter(csv_output, fieldnames=["Document ID", "Document Title", "Document File Name"])
            writer.writerow({
                "Document ID": document.id,
                "Document Title": document_meta_data.title,
                "Document File Name": document_meta_data.original_file_name,
            })
        counter += 1


# Function to translate certain metadata fields to Kannada & Hindi
async def translate_meta_data(jiva_library: Library, translator: GoogleTranslator):
    catalog = await jiva_library.catalog()
    with open("tools/docs_meta_data.csv", "r") as csv_input:
        reader = csv.DictReader(csv_input)
        counter = 1
        for row in reader:
            cat = row["Document ID"]
            print("\nFile Count:", counter)
            print("Document ID:", cat)
            meta_data = catalog[cat]
            title = meta_data.title
            legal_act_title = meta_data.extra_data["legal_act_title"]
            legal_ministry = meta_data.extra_data["legal_ministry"]
            kn_translated_title = await translator.translate_text(title, Language.EN, Language.KN)
            hi_translated_title = await translator.translate_text(title, Language.EN, Language.HI)
            kn_translated_legal_act_title = await translator.translate_text(legal_act_title, Language.EN, Language.KN)
            hi_translated_legal_act_title = await translator.translate_text(legal_act_title, Language.EN, Language.HI)
            if legal_ministry != "":
                kn_translated_legal_ministry = await translator.translate_text(legal_ministry, Language.EN, Language.KN)
                hi_translated_legal_ministry = await translator.translate_text(legal_ministry, Language.EN, Language.HI)
            else:
                kn_translated_legal_ministry = ""
                hi_translated_legal_ministry = ""
            with open("tools/translated_new_meta_data.csv", "a", newline="") as csv_output:
                writer = csv.DictWriter(csv_output, fieldnames=["Document ID", "Title", "Legal Act Title",
                                                                "Legal Ministry", "Title in Kannada", "Legal Act Title in Kannada",
                                                                "Legal Ministry in Kannada", "Title in Hindi", "Legal Act Title in Hindi",
                                                                "Legal Ministry in Hindi"])
                writer.writerow({
                    "Document ID": cat,
                    "Title": title,
                    "Legal Act Title": legal_act_title,
                    "Legal Ministry": legal_ministry,
                    "Title in Kannada": kn_translated_title,
                    "Legal Act Title in Kannada": kn_translated_legal_act_title,
                    "Legal Ministry in Kannada": kn_translated_legal_ministry,
                    "Title in Hindi": hi_translated_title,
                    "Legal Act Title in Hindi": hi_translated_legal_act_title,
                    "Legal Ministry in Hindi": hi_translated_legal_ministry
                })
            counter += 1


# Function to update translated metadata fields in DocumentMetaData object for each document and upload it to cloud storage
async def update_translated_metadata(jiva_library: Library):
    catalog = await jiva_library.catalog()
    with open("tools/translated_new_meta_data.csv", "r") as csv_input:
        reader = csv.DictReader(csv_input)
        counter = 1
        for row in reader:
            cat = row["Document ID"]
            print("\nFile Count:", counter)
            print("Document ID:", cat)
            document: Document = jiva_library.get_document(cat)
            meta_data = catalog[cat]
            meta_data.translated_data = {
                "title": {
                    "Kannada": row["Title in Kannada"],
                    "Hindi": row["Title in Hindi"]
                },
                "legal_act_title": {
                    "Kannada": row["Legal Act Title in Kannada"],
                    "Hindi": row["Legal Act Title in Hindi"]
                },
                "legal_ministry": {
                    "Kannada": row["Legal Ministry in Kannada"],
                    "Hindi": row["Legal Ministry in Hindi"]
                }
            }
            await document.write_metadata(meta_data)
            counter += 1


if __name__ == "__main__":
    load_dotenv()
    jiva_library = Library(id="jiva",
                           store=GoogleStorage(
                                bucket_name=os.environ["JIVA_LIBRARY_BUCKET"],
                                base_path=os.environ["JIVA_LIBRARY_PATH"]))
    # Run the below command once separately for converting given sheet to csv file
    # convert_google_sheets_meta_data_to_csv(required_sheet_name="Data_Anmol")
    # Run the below command once separately for uploading docs in given csv file
    # asyncio.run(act_uploading_process(jiva_library=jiva_library, csv_file_name="Data_Anmol.csv"))
    # Run the below command once separately for translating metadata
    # asyncio.run(translate_meta_data(jiva_library=jiva_library, translator=GoogleTranslator()))
    # Run the below command once separately to add translated fields to metadata
    asyncio.run(update_translated_metadata(jiva_library=jiva_library))
